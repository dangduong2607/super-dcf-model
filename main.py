from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil, os, zipfile, re
from tempfile import NamedTemporaryFile
from io import BytesIO

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])

def copy_sheet(source_sheet, target_wb, sheet_name):
    new_sheet = target_wb.create_sheet(sheet_name)
    for col in range(1, source_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        new_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
    for row in range(1, source_sheet.max_row + 1):
        new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
    for merged_range in source_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_range))
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.data_type == 'f':
                new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()
    return new_sheet

@app.post("/upload")
async def upload(consensus: UploadFile = File(...), profile: UploadFile = File(None)):
    consensus_path = "temp_consensus.xlsx"
    profile_path = None
    temp_file_path = None

    try:
        with open(consensus_path, "wb") as f:
            shutil.copyfileobj(consensus.file, f)
        if profile:
            profile_path = "temp_profile.xlsx"
            with open(profile_path, "wb") as f:
                shutil.copyfileobj(profile.file, f)

        # -------------------------------------------------------
        # STEP 1: Use openpyxl ONLY to process the user's consensus
        #         file (simple xlsx, no VBA/dynamic arrays).
        # -------------------------------------------------------
        consensus_wb = load_workbook(consensus_path)

        # Build a temporary xlsx containing only the user's extra sheets
        extra_wb_path = "/tmp/extra_sheets.xlsx"
        from openpyxl import Workbook
        extra_wb = Workbook()
        extra_wb.remove(extra_wb.active)  # remove default sheet

        for sheet_name in consensus_wb.sheetnames:
            if sheet_name == "DCF Model":
                continue
            copy_sheet(consensus_wb[sheet_name], extra_wb, sheet_name)

        extra_wb.save(extra_wb_path)

        # -------------------------------------------------------
        # STEP 2: Inject those sheets into the template using raw
        #         ZIP manipulation — never let openpyxl touch the
        #         template's own XML.
        # -------------------------------------------------------
        template_bytes = open("Template.xlsm", "rb").read()
        extra_bytes = open(extra_wb_path, "rb").read()

        output_buffer = BytesIO()

        with zipfile.ZipFile(BytesIO(template_bytes), 'r') as tmpl_zip, \
             zipfile.ZipFile(BytesIO(extra_bytes), 'r') as extra_zip, \
             zipfile.ZipFile(output_buffer, 'w', zipfile.ZIP_DEFLATED) as out_zip:

            # Copy ALL template parts verbatim
            for item in tmpl_zip.infolist():
                out_zip.writestr(item, tmpl_zip.read(item.filename))

            # Find how many sheets the template already has
            tmpl_wb_xml = tmpl_zip.read('xl/workbook.xml').decode('utf-8')
            existing_sheet_ids = [int(x) for x in re.findall(r'sheetId=\"(\d+)\"', tmpl_wb_xml)]
            max_sheet_id = max(existing_sheet_ids) if existing_sheet_ids else 1
            existing_rids = re.findall(r'r:id=\"(rId\d+)\"', tmpl_wb_xml)
            max_rid = max(int(r[3:]) for r in existing_rids) if existing_rids else 1

            # Read template's workbook rels
            rels_xml = tmpl_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            ct_xml = tmpl_zip.read('[Content_Types].xml').decode('utf-8')

            new_sheet_defs = ""
            new_rels = ""
            new_ct = ""
            sheet_counter = 0

            extra_wb_xml = extra_zip.read('xl/workbook.xml').decode('utf-8')
            extra_sheet_names = re.findall(r'name=\"([^\"]+)\"', extra_wb_xml)

            for i, sheet_name in enumerate(extra_sheet_names):
                sheet_counter += 1
                new_sheet_id = max_sheet_id + sheet_counter
                new_rid = f"rId{max_rid + sheet_counter}"
                internal_name = f"xl/worksheets/sheet_extra_{sheet_counter}.xml"

                # Copy sheet xml from extra workbook
                orig_sheet_xml = extra_zip.read(f'xl/worksheets/sheet{i+1}.xml')
                out_zip.writestr(internal_name, orig_sheet_xml)

                new_sheet_defs += f'<sheet name="{sheet_name}" sheetId="{new_sheet_id}" r:id="{new_rid}"/>'
                new_rels += f'<Relationship Id="{new_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet_extra_{sheet_counter}.xml"/>'
                new_ct += f'<Override PartName="/xl/worksheets/sheet_extra_{sheet_counter}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'

            # Patch workbook.xml to add new sheet references
            patched_wb = tmpl_wb_xml.replace('</sheets>', new_sheet_defs + '</sheets>')
            out_zip.writestr('xl/workbook.xml', patched_wb.encode('utf-8'))

            # Patch rels
            patched_rels = rels_xml.replace('</Relationships>', new_rels + '</Relationships>')
            out_zip.writestr('xl/_rels/workbook.xml.rels', patched_rels.encode('utf-8'))

            # Patch Content_Types
            patched_ct = ct_xml.replace('</Types>', new_ct + '</Types>')
            out_zip.writestr('[Content_Types].xml', patched_ct.encode('utf-8'))

        output_buffer.seek(0)

        with NamedTemporaryFile(delete=False, suffix=".xlsm") as temp_file:
            temp_file.write(output_buffer.read())
            temp_file_path = temp_file.name

        return StreamingResponse(
            open(temp_file_path, "rb"),
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            headers={"Content-Disposition": "attachment; filename=DCF_Model.xlsm"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        for path in [consensus_path, profile_path, temp_file_path, "/tmp/extra_sheets.xlsx"]:
            if path and os.path.exists(path):
                os.remove(path)
